import io
from typing import List
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
import tempfile
import os
import re
from datetime import datetime
from sqlalchemy import func, case, and_
from app import db
from app.models import Atencion, Tramite, Usuario, Area, Ticket
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer,
    Table, TableStyle, PageBreak, Image as RLImage
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER


class ReportService:

    ESTADO_FINALIZADO = "finalizado"
    ESTADO_CANCELADO = "cancelado"
    ESTADO_REASIGNADO = "reasignado"

    @staticmethod
    def _format_header_text(text):
        """Format header text: replace underscores with spaces and title case"""
        if not isinstance(text, str):
            return text
        return text.replace('_', ' ').title()

    @staticmethod
    def generar_reporte(
        fecha_inicio,
        fecha_fin,
        area_id=None,
        modo="tramites",
        exportar="excel",
        metricas_config=None
    ):
        """
        Genera reportes estadísticos de atenciones
        
        Args:
            fecha_inicio: Fecha inicio del periodo
            fecha_fin: Fecha fin del periodo
            area_id: ID del área (opcional)
            modo: 'tramites', 'usuarios', 'ambos'
            exportar: 'excel' o 'pdf'
            metricas_config: Dict con configuración de métricas a incluir
        """
        filtros = ReportService._build_filters(
            fecha_inicio,
            fecha_fin,
            area_id
        )

        if metricas_config is None:
            metricas_config = {
                'incluir_resumen': True,
                'incluir_estadisticas_base': True,
                'incluir_tiempos': True,
                'incluir_horas_pico': True,
                'incluir_horas_pico_semanal': True,
                'incluir_tabla_cruzada': False,
                'incluir_descripciones': False,
            }

        data = {}
        metadata = {
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'area_id': area_id,
            'modo': modo,
            'metricas_config': metricas_config
        }

        # Resumen general
        if metricas_config.get('incluir_resumen', True):
            data["resumen_general"] = ReportService._resumen_general(filtros, metricas_config)

        # Estadísticas base
        if metricas_config.get('incluir_estadisticas_base', True):
            if modo in ("tramites", "ambos"):
                data["tramites"] = ReportService._stats_por_tramite(filtros, metricas_config)
            if modo in ("usuarios", "ambos"):
                data["usuarios"] = ReportService._stats_por_usuario(filtros, metricas_config)

        # Descripciones de estados
        if metricas_config.get('incluir_descripciones', False):
            data["descripciones_general"] = ReportService._descripciones_general(filtros)
            
            if modo in ("tramites", "ambos"):
                data["descripciones_tramites"] = ReportService._descripciones_por_tramite(filtros)
            
            if modo in ("usuarios", "ambos"):
                data["descripciones_usuarios"] = ReportService._descripciones_por_usuario(filtros)

        # Horas pico
        if metricas_config.get('incluir_horas_pico', True):
            data["horas_pico"] = ReportService._horas_pico(filtros)
        
        if metricas_config.get('incluir_horas_pico_semanal', True):
            data["horas_pico_dia"] = ReportService._horas_pico_por_dia(filtros)

        if exportar == "excel":
            return ReportService._exportar_excel(data, metadata, filtros)

        return ReportService._exportar_pdf(data, metadata, filtros)

    @staticmethod
    def generar_reporte_admin_general(
        fecha_inicio,
        fecha_fin,
        area_ids:List[int]=None,
        exportar="excel",
        metricas_config=None
    ):
        """
        Genera reportes estadísticos para el administrador general
        
        Args:
            fecha_inicio: Fecha inicio del periodo
            fecha_fin: Fecha fin del periodo
            area_ids: Lista de IDs de áreas (opcional). Si None, incluye todas las áreas
            exportar: 'excel' o 'pdf'
        """
        filtros = ReportService._build_filters_admin_general(
            fecha_inicio,
            fecha_fin,
            area_ids
        )
        
        if metricas_config is None:
            metricas_config = {
                'incluir_resumen': True,
                'incluir_estadisticas_areas': True,
                'incluir_horas_pico': True,
                'incluir_horas_pico_semanal': True
            }

        data = {}
        metadata = {
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'area_ids': area_ids,
            'tipo_reporte': 'admin_general',
            'metricas_config': metricas_config
        }

        # Resumen general
        if metricas_config.get('incluir_resumen', True):
            data["resumen_general"] = ReportService._resumen_general_admin(filtros, fecha_inicio, fecha_fin)

        # Estadísticas por área
        if metricas_config.get('incluir_estadisticas_areas', True):
            data["estadisticas_areas"] = ReportService._stats_por_area(filtros)

        # Horas pico
        if metricas_config.get('incluir_horas_pico', True):
            data["horas_pico"] = ReportService._horas_pico_admin(filtros)
        
        # Horas pico por día de la semana
        if metricas_config.get('incluir_horas_pico_semanal', True):
            data["horas_pico_dia"] = ReportService._horas_pico_por_dia_admin(filtros)

        if exportar == "excel":
            return ReportService._exportar_excel_admin_general(data, metadata)

        return ReportService._exportar_pdf_admin_general(data, metadata)

    @staticmethod
    def _build_filters(fecha_inicio, fecha_fin, area_id):
        filtros = []
        if fecha_inicio:
            filtros.append(Atencion.hora_inicio >= fecha_inicio)
        if fecha_fin:
            filtros.append(Atencion.hora_inicio <= fecha_fin)
        if area_id:
            filtros.append(Usuario.area_id == area_id)
        return filtros

    @staticmethod
    def _build_filters_admin_general(fecha_inicio, fecha_fin, area_ids):
        """Construye filtros para el administrador general"""
        filtros = []
        if fecha_inicio:
            filtros.append(Atencion.hora_inicio >= fecha_inicio)
        if fecha_fin:
            filtros.append(Atencion.hora_inicio <= fecha_fin)
        if area_ids:
            filtros.append(Usuario.area_id.in_(area_ids))
        return filtros

    @staticmethod
    def _resumen_general_admin(filtros, fecha_inicio, fecha_fin):
        """Resumen general para admin que incluye tanto atenciones como tickets"""
        # Resumen de atenciones
        query_atenciones = (
            db.session.query(
                func.count(Atencion.id_atencion).label('total_atenciones'),
                func.sum(case((Atencion.estado == ReportService.ESTADO_FINALIZADO, 1), else_=0)).label('atenciones_finalizadas'),
                func.sum(case((Atencion.estado == ReportService.ESTADO_REASIGNADO, 1), else_=0)).label('atenciones_reasignadas'),
                func.sum(case((Atencion.estado == ReportService.ESTADO_CANCELADO, 1), else_=0)).label('atenciones_canceladas'),
            )
            .join(Usuario, Usuario.id_usuario == Atencion.id_usuario)
            .filter(and_(*filtros))
        )

        result_atenciones = query_atenciones.first()

        filtros_tickets = []
        if fecha_inicio:
            filtros_tickets.append(Ticket.fecha_hora >= fecha_inicio)
        if fecha_fin:
            filtros_tickets.append(Ticket.fecha_hora <= fecha_fin)
        
        # Agregar estado != activo
        filtros_tickets.append(Ticket.estado != 'activo')

        query_tickets = (
            db.session.query(
                func.count(Ticket.id_ticket).label('total_tickets'),
                func.sum(case((Ticket.estado == 'finalizado', 1), else_=0)).label('tickets_finalizados'),
                func.sum(case((Ticket.estado == 'cancelado', 1), else_=0)).label('tickets_cancelados'),
            )
            .filter(and_(*filtros_tickets))
        )

        result_tickets = query_tickets.first()

        resumen = {
            # Métricas de atenciones
            'total_atenciones': result_atenciones.total_atenciones or 0,
            'atenciones_finalizadas': result_atenciones.atenciones_finalizadas or 0,
            'atenciones_reasignadas': result_atenciones.atenciones_reasignadas or 0,
            'atenciones_canceladas': result_atenciones.atenciones_canceladas or 0,
            # Métricas de tickets
            'total_tickets': result_tickets.total_tickets or 0,
            'tickets_finalizados': result_tickets.tickets_finalizados or 0,
            'tickets_cancelados': result_tickets.tickets_cancelados or 0,
        }

        return resumen

    @staticmethod
    def _stats_por_area(filtros):
        """Estadísticas agrupadas por área"""
        query = (
            db.session.query(
                Area.id_area,
                Area.name.label('nombre_area'),
                func.count(Atencion.id_atencion).label('total_atenciones'),
                func.sum(case((Atencion.estado == ReportService.ESTADO_FINALIZADO, 1), else_=0)).label('finalizadas'),
                func.sum(case((Atencion.estado == ReportService.ESTADO_REASIGNADO, 1), else_=0)).label('reasignadas'),
                func.sum(case((Atencion.estado == ReportService.ESTADO_CANCELADO, 1), else_=0)).label('canceladas'),
            )
            .join(Usuario, Usuario.id_usuario == Atencion.id_usuario)
            .join(Area, Area.id_area == Usuario.area_id)
            .filter(and_(*filtros))
            .group_by(Area.id_area, Area.name)
        )

        df = pd.DataFrame(query.all(), columns=[
            'id_area', 'nombre_area', 'total_atenciones',
            'finalizadas', 'reasignadas', 'canceladas'
        ])

        if not df.empty:
            df = df.fillna(0)
        
        return df

    @staticmethod
    def _horas_pico_admin(filtros):
        """Horas pico para admin general"""
        query = (
            db.session.query(
                func.hour(Atencion.hora_inicio).label('hora'),
                func.count(Atencion.id_atencion).label('total')
            )
            .join(Usuario, Usuario.id_usuario == Atencion.id_usuario)
            .filter(and_(*filtros))
            .group_by('hora')
            .order_by('hora')
        )

        df = pd.DataFrame(query.all(), columns=['hora', 'total'])
        if not df.empty:
            df['hora'] = df['hora'].astype(int)
        return df

    @staticmethod
    def _horas_pico_por_dia_admin(filtros):
        """Horas pico por día de la semana para admin general"""
        query = (
            db.session.query(
                func.dayofweek(Atencion.hora_inicio).label('dia_semana'),
                func.hour(Atencion.hora_inicio).label('hora'),
                func.count(Atencion.id_atencion).label('total')
            )
            .join(Usuario, Usuario.id_usuario == Atencion.id_usuario)
            .filter(and_(*filtros))
            .group_by('dia_semana', 'hora')
        )

        df = pd.DataFrame(query.all(), columns=['dia_semana', 'hora', 'total'])
        if not df.empty:
            dias_map = {1: 'Domingo', 2: 'Lunes', 3: 'Martes', 4: 'Miércoles', 5: 'Jueves', 6: 'Viernes', 7: 'Sábado'}
            df['dia_nombre'] = df['dia_semana'].map(dias_map)
            df['hora'] = df['hora'].astype(int)
        return df

    @staticmethod
    def _exportar_excel_admin_general(data, metadata):
        """Exporta el reporte del admin general a Excel"""
        output = io.BytesIO()
        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Resumen General
        if 'resumen_general' in data and data['resumen_general']:
            ws = wb.create_sheet('Resumen General')
            ws['A1'] = 'REPORTE GENERAL DEL SISTEMA'
            ws['A1'].font = Font(bold=True, size=14)
            
            row = 3
            ws[f'A{row}'] = 'ATENCIONES'
            ws[f'A{row}'].font = Font(bold=True, size=12, color='366092')
            row += 1
            
            atenciones_keys = ['total_atenciones', 'atenciones_finalizadas', 'atenciones_reasignadas', 'atenciones_canceladas']
            for key in atenciones_keys:
                ws[f'A{row}'] = ReportService._format_header_text(key)
                ws[f'B{row}'] = data['resumen_general'].get(key, 0)
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            row += 1
            ws[f'A{row}'] = 'TICKETS'
            ws[f'A{row}'].font = Font(bold=True, size=12, color='366092')
            row += 1
            
            tickets_keys = ['total_tickets', 'tickets_finalizados', 'tickets_cancelados']
            for key in tickets_keys:
                ws[f'A{row}'] = ReportService._format_header_text(key)
                ws[f'B{row}'] = data['resumen_general'].get(key, 0)
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

        # Estadísticas por Área
        if 'estadisticas_areas' in data and not data['estadisticas_areas'].empty:
            ws = wb.create_sheet('Estadísticas por Área')
            ReportService._write_dataframe_to_sheet(ws, data['estadisticas_areas'], header_font, header_fill, border)

        # Horas Pico
        if 'horas_pico' in data and not data['horas_pico'].empty:
            ws = wb.create_sheet('Horas Pico')
            ReportService._write_dataframe_to_sheet(ws, data['horas_pico'], header_font, header_fill, border)

        # Horas Pico por Día
        if 'horas_pico_dia' in data and not data['horas_pico_dia'].empty:
            ws = wb.create_sheet('Horas Pico por Día')
            ReportService._write_dataframe_to_sheet(ws, data['horas_pico_dia'], header_font, header_fill, border)

        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def _exportar_pdf_admin_general(data, metadata):
        """Exporta el reporte del admin general a PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=50, bottomMargin=50)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=13,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12
        )

        temp_image_paths = []

        try:
            elements.append(Paragraph("Reporte General del Sistema", title_style))
            elements.append(Spacer(1, 0.1*inch))

            def format_date_display(dt):
                if dt is None:
                    return 'N/A'
                
                if isinstance(dt, str):
                    try:
                        if '.' in dt:
                            dt = dt.split('.')[0]
                        dt = datetime.fromisoformat(dt.replace('Z', '+00:00'))
                    except:
                        return dt
                
                if isinstance(dt, datetime):
                    if dt.hour == 0 and dt.minute == 0 and dt.second == 0:
                        return dt.strftime('%Y-%m-%d')
                    return dt.strftime('%Y-%m-%d %H:%M:%S')
                
                return str(dt)

            fecha_inicio_str = format_date_display(metadata.get('fecha_inicio'))
            fecha_fin_str = format_date_display(metadata.get('fecha_fin'))
            
            elements.append(Paragraph(f"Periodo: {fecha_inicio_str} - {fecha_fin_str}", styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))

            # Resumen General
            if 'resumen_general' in data and data['resumen_general']:
                elements.append(Paragraph("Resumen General", heading_style))
                resumen = data['resumen_general']
                
                resumen_data = [
                    ['Métrica', 'Valor'],
                    ['', ''],
                    ['ATENCIONES', ''],
                    ['Total de Atenciones', f"{resumen.get('total_atenciones', 0):,}"],
                    ['Atenciones Finalizadas', f"{resumen.get('atenciones_finalizadas', 0):,}"],
                    ['Atenciones Reasignadas', f"{resumen.get('atenciones_reasignadas', 0):,}"],
                    ['Atenciones Canceladas', f"{resumen.get('atenciones_canceladas', 0):,}"],
                    ['', ''],
                    ['TICKETS', ''],
                    ['Total de Tickets', f"{resumen.get('total_tickets', 0):,}"],
                    ['Tickets Finalizados', f"{resumen.get('tickets_finalizados', 0):,}"],
                    ['Tickets Cancelados', f"{resumen.get('tickets_cancelados', 0):,}"],
                ]
                
                table = Table(resumen_data, colWidths=[3*inch, 2*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
                    ('FONTNAME', (0, 8), (-1, 8), 'Helvetica-Bold'),
                    ('TEXTCOLOR', (0, 2), (-1, 2), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 8), (-1, 8), colors.HexColor('#366092')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(table)
                elements.append(Spacer(1, 0.3*inch))

                # Gráfica de atenciones
                total_atenciones = sum([
                    resumen.get('atenciones_finalizadas', 0),
                    resumen.get('atenciones_reasignadas', 0),
                    resumen.get('atenciones_canceladas', 0)
                ])
                if total_atenciones > 0:
                    img_path = ReportService._crear_grafica_estados_admin(resumen)
                    if img_path:
                        temp_image_paths.append(img_path)
                        elements.append(Paragraph("Distribución de Estados de Atenciones", heading_style))
                        img = RLImage(img_path, width=4.0*inch, height=3.5*inch)
                        elements.append(img)
                        elements.append(Spacer(1, 0.5*inch))

                # Gráfica de tickets
                total_tickets = sum([
                    resumen.get('tickets_finalizados', 0),
                    resumen.get('tickets_cancelados', 0)
                ])
                if total_tickets > 0:
                    img_path = ReportService._crear_grafica_tickets(resumen)
                    if img_path:
                        temp_image_paths.append(img_path)
                        elements.append(Paragraph("Distribución de Estados de Tickets", heading_style))
                        img = RLImage(img_path, width=4.0*inch, height=3.5*inch)
                        elements.append(img)
                        elements.append(Spacer(1, 0.3*inch))

            # Estadísticas por Área
            if 'estadisticas_areas' in data and not data['estadisticas_areas'].empty:
                elements.append(PageBreak())
                elements.append(Paragraph("Estadísticas por Área", heading_style))
                
                df_areas = data['estadisticas_areas'].copy()
                df_areas['nombre_area'] = df_areas['nombre_area'].astype(str).apply(
                    lambda x: (x[:35] + '...') if len(x) > 35 else x
                )
                
                columnas_formateadas = [ReportService._format_header_text(col) for col in df_areas.columns]
                table_data = [columnas_formateadas] + df_areas.values.tolist()
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
                elements.append(Spacer(1, 0.3*inch))
                
                if df_areas['total_atenciones'].sum() > 0:
                    img_path = ReportService._crear_grafica_areas(data['estadisticas_areas'])
                    if img_path:
                        temp_image_paths.append(img_path)
                        elements.append(Paragraph("Comparación por Área", heading_style))
                        img = RLImage(img_path, width=5.0*inch, height=3.5*inch)
                        elements.append(img)
                        elements.append(Spacer(1, 0.3*inch))

            # Horas Pico
            if 'horas_pico' in data and not data['horas_pico'].empty and data['horas_pico']['total'].sum() > 0:
                elements.append(PageBreak())
                elements.append(Paragraph("Distribución de Atenciones por Hora", heading_style))
                img_path = ReportService._crear_grafica_horas_pico(data['horas_pico'])
                if img_path:
                    temp_image_paths.append(img_path)
                    img = RLImage(img_path, width=5.5*inch, height=3.0*inch)
                    elements.append(img)
                    elements.append(Spacer(1, 0.3*inch))

            # Patrón Semanal
            if 'horas_pico_dia' in data and not data['horas_pico_dia'].empty and data['horas_pico_dia']['total'].sum() > 0:
                elements.append(PageBreak())
                elements.append(Paragraph("Patrón Semanal de Atenciones", heading_style))
                img_path = ReportService._crear_heatmap_horas_dia(data['horas_pico_dia'])
                if img_path:
                    temp_image_paths.append(img_path)
                    img = RLImage(img_path, width=5.5*inch, height=3.5*inch)
                    elements.append(img)

            doc.build(elements)

        finally:
            for path in temp_image_paths:
                try:
                    os.unlink(path)
                except Exception as e:
                    print(f"Advertencia: No se pudo eliminar archivo temporal {path}: {e}")

        buffer.seek(0)
        return buffer

    @staticmethod
    def _crear_grafica_estados_admin(resumen):
        """Crea gráfica de estados de atenciones para admin"""
        try:
            estados = ['Finalizadas', 'Reasignadas', 'Canceladas']
            valores = [
                resumen.get('atenciones_finalizadas', 0),
                resumen.get('atenciones_reasignadas', 0),
                resumen.get('atenciones_canceladas', 0)
            ]
            if sum(valores) == 0:
                return None
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                fig, ax = plt.subplots(figsize=(6, 6))
                colors_pie = ['#2ecc71', '#f39c12', '#e74c3c']
                ax.pie(valores, labels=estados, autopct='%1.1f%%', startangle=90, colors=colors_pie)
                ax.set_title('Distribución de Estados de Atenciones', fontsize=12, fontweight='bold', pad=20)
                plt.tight_layout(pad=2.0)
                plt.savefig(tmp.name, dpi=150, bbox_inches='tight')
                plt.close()
                return tmp.name
        except Exception as e:
            print(f"Error creando gráfica de estados: {e}")
            return None

    @staticmethod
    def _crear_grafica_tickets(resumen):
        """Crea gráfica de estados de tickets"""
        try:
            estados = ['Finalizados', 'Cancelados']
            valores = [
                resumen.get('tickets_finalizados', 0),
                resumen.get('tickets_cancelados', 0)
            ]
            if sum(valores) == 0:
                return None
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                fig, ax = plt.subplots(figsize=(6, 6))
                colors_pie = ['#3498db', '#e74c3c']
                ax.pie(valores, labels=estados, autopct='%1.1f%%', startangle=90, colors=colors_pie)
                ax.set_title('Distribución de Estados de Tickets', fontsize=12, fontweight='bold', pad=20)
                plt.tight_layout(pad=2.0)
                plt.savefig(tmp.name, dpi=150, bbox_inches='tight')
                plt.close()
                return tmp.name
        except Exception as e:
            print(f"Error creando gráfica de tickets: {e}")
            return None

    @staticmethod
    def _crear_grafica_areas(df):
        """Crea gráfica comparativa por áreas"""
        try:
            if df.empty or df['total_atenciones'].sum() == 0:
                return None
            df_sorted = df.nlargest(10, 'total_atenciones').copy()
            df_sorted['nombre_area'] = df_sorted['nombre_area'].astype(str).apply(
                lambda x: (x[:40] + '...') if len(x) > 40 else x
            )
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                fig, ax = plt.subplots(figsize=(7, 5))
                bars = ax.barh(df_sorted['nombre_area'], df_sorted['total_atenciones'], color='#3498db')
                ax.set_xlabel('Total de Atenciones', fontsize=10)
                ax.set_title('Atenciones por Área', fontsize=12, fontweight='bold', pad=15)
                ax.invert_yaxis()
                ax.tick_params(axis='y', labelsize=8)
                ax.set_xlim(0, df_sorted['total_atenciones'].max() * 1.1)
                for bar in bars:
                    width = bar.get_width()
                    ax.text(width + 0.5, bar.get_y() + bar.get_height()/2, str(int(width)),
                            va='center', ha='left', fontsize=8)
                plt.tight_layout(pad=2.0)
                plt.savefig(tmp.name, dpi=150, bbox_inches='tight')
                plt.close()
                return tmp.name
        except Exception as e:
            print(f"Error creando gráfica de áreas: {e}")
            return None

    # ===== MÉTODOS EXISTENTES (sin modificar) =====

    @staticmethod
    def _stats_por_tramite(filtros, metricas_config):
        """
        Ahora usa Atencion.id_tramite directamente para obtener el trámite original
        """
        incluir_tiempos = metricas_config.get('incluir_tiempos', True)
        
        # Columnas base
        columnas_query = [
            Atencion.id_tramite,
            Tramite.name.label('nombre_tramite'),
            func.count(Atencion.id_atencion).label('total_atenciones'),
            func.sum(case((Atencion.estado == ReportService.ESTADO_FINALIZADO, 1), else_=0)).label('finalizadas'),
            func.sum(case((Atencion.estado == ReportService.ESTADO_REASIGNADO, 1), else_=0)).label('reasignadas'),
            func.sum(case((Atencion.estado == ReportService.ESTADO_CANCELADO, 1), else_=0)).label('canceladas'),
        ]
        
        columnas_df = [
            'id_tramite', 'nombre_tramite', 'total_atenciones', 
            'finalizadas', 'reasignadas', 'canceladas'
        ]
        
        # Agregar columna de tiempo solo si está habilitada
        if incluir_tiempos:
            columnas_query.append(
                func.avg(func.timestampdiff(db.text("SECOND"), Atencion.hora_inicio, Atencion.hora_fin)).label('tiempo_promedio_segundos')
            )
            columnas_df.append('tiempo_promedio_segundos')
        
        query = (
            db.session.query(*columnas_query)
            .join(Tramite, Tramite.id_tramite == Atencion.id_tramite)
            .join(Usuario, Usuario.id_usuario == Atencion.id_usuario)
            .filter(and_(*filtros))
            .group_by(Atencion.id_tramite, Tramite.name)
        )

        df = pd.DataFrame(query.all(), columns=columnas_df)

        if not df.empty:
            if incluir_tiempos:
                df['tiempo_promedio_minutos'] = (df['tiempo_promedio_segundos'] / 60).round(2)
                df = df.drop(columns=['tiempo_promedio_segundos'])
            df = df.fillna(0)
        
        return df

    @staticmethod
    def _stats_por_usuario(filtros, metricas_config):
        incluir_tiempos = metricas_config.get('incluir_tiempos', True)
        
        # Columnas base
        columnas_query = [
            Atencion.id_usuario,
            Usuario.username.label('username'),
            func.concat(Usuario.nombre, ' ', Usuario.ap_paterno, ' ', func.coalesce(Usuario.ap_materno, '')).label('nombre_completo'),
            func.count(Atencion.id_atencion).label('total_atenciones'),
            func.sum(case((Atencion.estado == ReportService.ESTADO_FINALIZADO, 1), else_=0)).label('finalizadas'),
            func.sum(case((Atencion.estado == ReportService.ESTADO_REASIGNADO, 1), else_=0)).label('reasignadas'),
            func.sum(case((Atencion.estado == ReportService.ESTADO_CANCELADO, 1), else_=0)).label('canceladas'),
        ]
        
        columnas_df = [
            'id_usuario', 'username', 'nombre_completo', 'total_atenciones',
            'finalizadas', 'reasignadas', 'canceladas'
        ]
        
        # Agregar columna de tiempo solo si está habilitada
        if incluir_tiempos:
            columnas_query.append(
                func.avg(func.timestampdiff(db.text("SECOND"), Atencion.hora_inicio, Atencion.hora_fin)).label('tiempo_promedio_segundos')
            )
            columnas_df.append('tiempo_promedio_segundos')
        
        query = (
            db.session.query(*columnas_query)
            .join(Usuario, Usuario.id_usuario == Atencion.id_usuario)
            .filter(and_(*filtros))
            .group_by(Atencion.id_usuario, Usuario.username, 'nombre_completo')
        )

        df = pd.DataFrame(query.all(), columns=columnas_df)

        if not df.empty:
            if incluir_tiempos:
                df['tiempo_promedio_minutos'] = (df['tiempo_promedio_segundos'] / 60).round(2)
                df = df.drop(columns=['tiempo_promedio_segundos'])
            df = df.fillna(0)
            df['nombre_completo'] = df['nombre_completo'].str.strip()
        
        return df

    @staticmethod
    def _descripciones_general(filtros):
        """
        Agrupa las descripciones de estado por cada estado (cancelado, reasignado, finalizado)
        """
        query = (
            db.session.query(
                Atencion.estado,
                Atencion.descripcion_estado,
                func.count(Atencion.id_atencion).label('total')
            )
            .join(Usuario, Usuario.id_usuario == Atencion.id_usuario)
            .filter(and_(*filtros))
            .filter(Atencion.descripcion_estado.isnot(None))
            .filter(Atencion.descripcion_estado != '')
            .group_by(Atencion.estado, Atencion.descripcion_estado)
            .order_by(Atencion.estado, func.count(Atencion.id_atencion).desc())
        )
        
        df = pd.DataFrame(query.all(), columns=['estado', 'descripcion_estado', 'total'])
        
        # Crear un diccionario con un DataFrame por cada estado
        result = {}
        if not df.empty:
            for estado in df['estado'].unique():
                df_estado = df[df['estado'] == estado][['descripcion_estado', 'total']].copy()
                df_estado.columns = ['motivo', 'total']
                result[estado] = df_estado
        
        return result

    @staticmethod
    def _descripciones_por_tramite(filtros):
        """
        Agrupa las descripciones de estado por trámite y estado
        """
        query = (
            db.session.query(
                Atencion.id_tramite,
                Tramite.name.label('nombre_tramite'),
                Atencion.estado,
                Atencion.descripcion_estado,
                func.count(Atencion.id_atencion).label('total')
            )
            .join(Tramite, Tramite.id_tramite == Atencion.id_tramite)
            .join(Usuario, Usuario.id_usuario == Atencion.id_usuario)
            .filter(and_(*filtros))
            .filter(Atencion.descripcion_estado.isnot(None))
            .filter(Atencion.descripcion_estado != '')
            .group_by(Atencion.id_tramite, Tramite.name, Atencion.estado, Atencion.descripcion_estado)
            .order_by(Tramite.name, Atencion.estado, func.count(Atencion.id_atencion).desc())
        )
        
        df = pd.DataFrame(query.all(), columns=['id_tramite', 'nombre_tramite', 'estado', 'descripcion_estado', 'total'])
        
        # Crear un diccionario con DataFrames por estado
        result = {}
        if not df.empty:
            for estado in df['estado'].unique():
                df_estado = df[df['estado'] == estado][['nombre_tramite', 'descripcion_estado', 'total']].copy()
                df_estado.columns = ['tramite', 'motivo', 'total']
                result[estado] = df_estado
        
        return result

    @staticmethod
    def _descripciones_por_usuario(filtros):
        """
        Agrupa las descripciones de estado por usuario y estado
        """
        query = (
            db.session.query(
                Atencion.id_usuario,
                func.concat(Usuario.nombre, ' ', Usuario.ap_paterno, ' ', func.coalesce(Usuario.ap_materno, '')).label('nombre_completo'),
                Atencion.estado,
                Atencion.descripcion_estado,
                func.count(Atencion.id_atencion).label('total')
            )
            .join(Usuario, Usuario.id_usuario == Atencion.id_usuario)
            .filter(and_(*filtros))
            .filter(Atencion.descripcion_estado.isnot(None))
            .filter(Atencion.descripcion_estado != '')
            .group_by(Atencion.id_usuario, 'nombre_completo', Atencion.estado, Atencion.descripcion_estado)
            .order_by('nombre_completo', Atencion.estado, func.count(Atencion.id_atencion).desc())
        )
        
        df = pd.DataFrame(query.all(), columns=['id_usuario', 'nombre_completo', 'estado', 'descripcion_estado', 'total'])
        
        # Crear un diccionario con DataFrames por estado
        result = {}
        if not df.empty:
            df['nombre_completo'] = df['nombre_completo'].str.strip()
            for estado in df['estado'].unique():
                df_estado = df[df['estado'] == estado][['nombre_completo', 'descripcion_estado', 'total']].copy()
                df_estado.columns = ['usuario', 'motivo', 'total']
                result[estado] = df_estado
        
        return result

    @staticmethod
    def _tabla_cruzada_global(filtros):
        """Matriz completa empleado x trámite (solo para Excel)"""
        query = (
            db.session.query(
                func.concat(Usuario.nombre, ' ', Usuario.ap_paterno, ' ', func.coalesce(Usuario.ap_materno, '')).label('empleado'),
                Tramite.name.label('tramite'),
                func.count(Atencion.id_atencion).label('total')
            )
            .join(Usuario, Usuario.id_usuario == Atencion.id_usuario)
            .join(Tramite, Tramite.id_tramite == Atencion.id_tramite)
            .filter(and_(*filtros))
            .group_by('empleado', Tramite.name)
        )

        df = pd.DataFrame(query.all(), columns=['empleado', 'tramite', 'total'])
        if df.empty:
            return pd.DataFrame()
        df['empleado'] = df['empleado'].str.strip()
        tabla = pd.pivot_table(df, values='total', index='empleado', columns='tramite', fill_value=0, aggfunc='sum')
        return tabla

    @staticmethod
    def _tramites_por_empleado(filtros):
        """Devuelve dict: {empleado: DataFrame(tramite, total)}"""
        query = (
            db.session.query(
                func.concat(Usuario.nombre, ' ', Usuario.ap_paterno, ' ', func.coalesce(Usuario.ap_materno, '')).label('empleado'),
                Tramite.name.label('tramite'),
                func.count(Atencion.id_atencion).label('total')
            )
            .join(Usuario, Usuario.id_usuario == Atencion.id_usuario)
            .join(Tramite, Tramite.id_tramite == Atencion.id_tramite)
            .filter(and_(*filtros))
            .group_by('empleado', Tramite.name)
            .order_by('empleado', Tramite.name)
        )
        
        df = pd.DataFrame(query.all(), columns=['empleado', 'tramite', 'total'])
        if df.empty:
            return {}
        
        empleado_dict = {}
        for empleado, group in df.groupby('empleado'):
            empleado_dict[empleado] = group[['tramite', 'total']].sort_values('total', ascending=False)
        return empleado_dict

    @staticmethod
    def _horas_pico(filtros):
        query = (
            db.session.query(
                func.hour(Atencion.hora_inicio).label('hora'),
                func.count(Atencion.id_atencion).label('total')
            )
            .join(Usuario, Usuario.id_usuario == Atencion.id_usuario)
            .filter(and_(*filtros))
            .group_by('hora')
            .order_by('hora')
        )

        df = pd.DataFrame(query.all(), columns=['hora', 'total'])
        if not df.empty:
            df['hora'] = df['hora'].astype(int)
        return df

    @staticmethod
    def _horas_pico_por_dia(filtros):
        query = (
            db.session.query(
                func.dayofweek(Atencion.hora_inicio).label('dia_semana'),
                func.hour(Atencion.hora_inicio).label('hora'),
                func.count(Atencion.id_atencion).label('total')
            )
            .join(Usuario, Usuario.id_usuario == Atencion.id_usuario)
            .filter(and_(*filtros))
            .group_by('dia_semana', 'hora')
        )

        df = pd.DataFrame(query.all(), columns=['dia_semana', 'hora', 'total'])
        if not df.empty:
            dias_map = {1: 'Domingo', 2: 'Lunes', 3: 'Martes', 4: 'Miércoles', 5: 'Jueves', 6: 'Viernes', 7: 'Sábado'}
            df['dia_nombre'] = df['dia_semana'].map(dias_map)
            df['hora'] = df['hora'].astype(int)
        return df

    @staticmethod
    def _resumen_general(filtros, metricas_config):
        incluir_tiempos = metricas_config.get('incluir_tiempos', True)
        
        columnas_query = [
            func.count(Atencion.id_atencion).label('total_atenciones'),
            func.sum(case((Atencion.estado == ReportService.ESTADO_FINALIZADO, 1), else_=0)).label('finalizadas'),
            func.sum(case((Atencion.estado == ReportService.ESTADO_REASIGNADO, 1), else_=0)).label('reasignadas'),
            func.sum(case((Atencion.estado == ReportService.ESTADO_CANCELADO, 1), else_=0)).label('canceladas'),
            func.count(func.distinct(Atencion.id_usuario)).label('usuarios_activos'),
            func.count(func.distinct(Atencion.id_tramite)).label('tramites_distintos')
        ]
        
        if incluir_tiempos:
            columnas_query.insert(4, func.avg(func.timestampdiff(db.text("SECOND"), Atencion.hora_inicio, Atencion.hora_fin)).label('tiempo_promedio_segundos'))
        
        query = (
            db.session.query(*columnas_query)
            .join(Usuario, Usuario.id_usuario == Atencion.id_usuario)
            .filter(and_(*filtros))
        )

        result = query.first()
        if result:
            resumen = {
                'total_atenciones': result.total_atenciones or 0,
                'finalizadas': result.finalizadas or 0,
                'reasignadas': result.reasignadas or 0,
                'canceladas': result.canceladas or 0,
                'usuarios_activos': result.usuarios_activos or 0,
                'tramites_distintos': result.tramites_distintos or 0
            }
            
            if incluir_tiempos:
                resumen['tiempo_promedio_minutos'] = round((result.tiempo_promedio_segundos or 0) / 60, 2)
            
            return resumen
        
        return {}

    @staticmethod
    def _exportar_excel(data, metadata, filtros):
        output = io.BytesIO()
        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Resumen
        if 'resumen_general' in data and data['resumen_general']:
            ws = wb.create_sheet('Resumen General')
            ws['A1'] = 'RESUMEN GENERAL'
            ws['A1'].font = Font(bold=True, size=14)
            row = 3
            for key, value in data['resumen_general'].items():
                ws[f'A{row}'] = ReportService._format_header_text(key)
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

        # Tramites
        if 'tramites' in data and not data['tramites'].empty:
            ws = wb.create_sheet('Estadísticas por Trámite')
            ReportService._write_dataframe_to_sheet(ws, data['tramites'], header_font, header_fill, border)

        # Usuarios
        if 'usuarios' in data and not data['usuarios'].empty:
            ws = wb.create_sheet('Estadísticas por Usuario')
            ReportService._write_dataframe_to_sheet(ws, data['usuarios'], header_font, header_fill, border)

        # Descripciones generales
        if 'descripciones_general' in data and data['descripciones_general']:
            for estado, df_desc in data['descripciones_general'].items():
                if not df_desc.empty:
                    sheet_name = f'Motivos {estado.title()}'
                    ws = wb.create_sheet(sheet_name)
                    ReportService._write_dataframe_to_sheet(ws, df_desc, header_font, header_fill, border)

        # Descripciones por trámite
        if 'descripciones_tramites' in data and data['descripciones_tramites']:
            for estado, df_desc in data['descripciones_tramites'].items():
                if not df_desc.empty:
                    sheet_name = f'Motivos {estado.title()} x Trámite'
                    ws = wb.create_sheet(sheet_name)
                    ReportService._write_dataframe_to_sheet(ws, df_desc, header_font, header_fill, border)

        # Descripciones por usuario
        if 'descripciones_usuarios' in data and data['descripciones_usuarios']:
            for estado, df_desc in data['descripciones_usuarios'].items():
                if not df_desc.empty:
                    sheet_name = f'Motivos {estado.title()} x Usuario'
                    ws = wb.create_sheet(sheet_name)
                    ReportService._write_dataframe_to_sheet(ws, df_desc, header_font, header_fill, border)

        # Matriz global
        if metadata.get('metricas_config', {}).get('incluir_tabla_cruzada', False):
            tabla_global = ReportService._tabla_cruzada_global(filtros)
            if not tabla_global.empty:
                ws = wb.create_sheet('Matriz Empleado-Trámite')
                ReportService._write_dataframe_to_sheet(ws, tabla_global, header_font, header_fill, border, include_index=True)

        # Hojas individuales por empleado
        tramites_por_emp = ReportService._tramites_por_empleado(filtros)
        for empleado, df_emp in tramites_por_emp.items():
            sheet_name = re.sub(r'[\\/*?:"<>|]', "_", empleado)[:30]
            ws_emp = wb.create_sheet(sheet_name)
            ReportService._write_dataframe_to_sheet(ws_emp, df_emp, header_font, header_fill, border)

        # Horas pico
        if 'horas_pico' in data and not data['horas_pico'].empty:
            ws = wb.create_sheet('Horas Pico')
            ReportService._write_dataframe_to_sheet(ws, data['horas_pico'], header_font, header_fill, border)

        if 'horas_pico_dia' in data and not data['horas_pico_dia'].empty:
            ws = wb.create_sheet('Horas Pico por Día')
            ReportService._write_dataframe_to_sheet(ws, data['horas_pico_dia'], header_font, header_fill, border)

        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def _write_dataframe_to_sheet(ws, df, header_font, header_fill, border, include_index=False):
        formatted_columns = [ReportService._format_header_text(col) for col in df.columns]
        df_formatted = df.copy()
        df_formatted.columns = formatted_columns
        
        for r_idx, row in enumerate(dataframe_to_rows(df_formatted, index=include_index, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                if r_idx == 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def _exportar_pdf(data, metadata, filtros):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=50, bottomMargin=50)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=13,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12
        )

        temp_image_paths = []
        metricas_config = metadata.get('metricas_config', {})
        incluir_tiempos = metricas_config.get('incluir_tiempos', True)

        try:
            elements.append(Paragraph("Reporte Estadístico de Atenciones", title_style))
            elements.append(Spacer(1, 0.1*inch))

            def format_date_display(dt):
                if dt is None:
                    return 'N/A'
                
                if isinstance(dt, str):
                    try:
                        if '.' in dt:
                            dt = dt.split('.')[0]
                        dt = datetime.fromisoformat(dt.replace('Z', '+00:00'))
                    except:
                        return dt
                
                if isinstance(dt, datetime):
                    if dt.hour == 0 and dt.minute == 0 and dt.second == 0:
                        return dt.strftime('%Y-%m-%d')
                    return dt.strftime('%Y-%m-%d %H:%M:%S')
                
                return str(dt)

            fecha_inicio_str = format_date_display(metadata.get('fecha_inicio'))
            fecha_fin_str = format_date_display(metadata.get('fecha_fin'))
            
            # Agregar información del area
            area_id = metadata.get('area_id')
            if area_id:
                area = db.session.query(Area).filter_by(id_area=area_id).first()
                if area:
                    elements.append(Paragraph(f"Área: {area.name}", styles['Heading3']))
                    elements.append(Spacer(1, 0.1*inch))
            
            elements.append(Paragraph(f"Periodo: {fecha_inicio_str} - {fecha_fin_str}", styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))

            # Resumen General
            if 'resumen_general' in data and data['resumen_general']:
                elements.append(Paragraph("Resumen General", heading_style))
                resumen = data['resumen_general']
                
                resumen_data = [
                    ['Métrica', 'Valor'],
                    ['Total de Atenciones', f"{resumen.get('total_atenciones', 0):,}"],
                    ['Atendidas', f"{resumen.get('finalizadas', 0):,}"],
                    ['Reasignadas', f"{resumen.get('reasignadas', 0):,}"],
                    ['Canceladas', f"{resumen.get('canceladas', 0):,}"],
                ]
                
                if incluir_tiempos:
                    resumen_data.append(['Tiempo Promedio (min)', f"{resumen.get('tiempo_promedio_minutos', 0):.2f}"])
                
                resumen_data.extend([
                    ['Usuarios Activos', f"{resumen.get('usuarios_activos', 0):,}"],
                    ['Trámites Distintos', f"{resumen.get('tramites_distintos', 0):,}"]
                ])
                
                table = Table(resumen_data, colWidths=[3*inch, 2*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(table)
                elements.append(Spacer(1, 0.3*inch))

                total_estados = sum([resumen.get('finalizadas',0), resumen.get('reasignadas',0), resumen.get('canceladas',0)])
                if total_estados > 0:
                    img_path = ReportService._crear_grafica_estados(resumen)
                    if img_path:
                        temp_image_paths.append(img_path)
                        elements.append(Paragraph("Distribución de Estados de Atenciones", heading_style))
                        img = RLImage(img_path, width=4.0*inch, height=3.5*inch)
                        elements.append(img)
                        elements.append(Spacer(1, 0.3*inch))

            # Estadísticas por Trámite
            if 'tramites' in data and not data['tramites'].empty:
                elements.append(PageBreak())
                elements.append(Paragraph("Estadísticas por Trámite", heading_style))
                
                # Seleccionar columnas según si incluir_tiempos está activo
                if incluir_tiempos:
                    columnas_a_mostrar = ['nombre_tramite', 'total_atenciones', 'finalizadas', 
                                         'reasignadas', 'canceladas', 'tiempo_promedio_minutos']
                else:
                    columnas_a_mostrar = ['nombre_tramite', 'total_atenciones', 'finalizadas', 
                                         'reasignadas', 'canceladas']
                
                df_tramites = data['tramites'][columnas_a_mostrar].head(10).copy()
                df_tramites['nombre_tramite'] = df_tramites['nombre_tramite'].astype(str).apply(
                    lambda x: (x[:35] + '...') if len(x) > 35 else x
                )
                
                columnas_formateadas = [ReportService._format_header_text(col) for col in df_tramites.columns]
                table_data = [columnas_formateadas] + df_tramites.values.tolist()
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
                elements.append(Spacer(1, 0.3*inch))
                
                if df_tramites['total_atenciones'].sum() > 0:
                    img_path = ReportService._crear_grafica_top_tramites(data['tramites'])
                    if img_path:
                        temp_image_paths.append(img_path)
                        elements.append(Paragraph("Trámites con Mayor Demanda", heading_style))
                        img = RLImage(img_path, width=5.0*inch, height=3.5*inch)
                        elements.append(img)
                        elements.append(Spacer(1, 0.3*inch))

            # Estadísticas por Empleado
            if 'usuarios' in data and not data['usuarios'].empty:
                elements.append(PageBreak())
                elements.append(Paragraph("Estadísticas por Empleado", heading_style))
                
                # Seleccionar columnas según si incluir_tiempos está activo
                if incluir_tiempos:
                    columnas_a_mostrar = ['nombre_completo', 'total_atenciones', 'finalizadas',
                                         'reasignadas', 'canceladas', 'tiempo_promedio_minutos']
                else:
                    columnas_a_mostrar = ['nombre_completo', 'total_atenciones', 'finalizadas',
                                         'reasignadas', 'canceladas']
                
                df_usuarios = data['usuarios'][columnas_a_mostrar].head(10).copy()
                df_usuarios['nombre_completo'] = df_usuarios['nombre_completo'].astype(str).apply(
                    lambda x: (x[:30] + '...') if len(x) > 30 else x
                )
                
                columnas_formateadas = [ReportService._format_header_text(col) for col in df_usuarios.columns]
                table_data = [columnas_formateadas] + df_usuarios.values.tolist()
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
                elements.append(Spacer(1, 0.3*inch))
                
                if df_usuarios['total_atenciones'].sum() > 0:
                    img_path = ReportService._crear_grafica_top_empleados(data['usuarios'])
                    if img_path:
                        temp_image_paths.append(img_path)
                        elements.append(Paragraph("Empleados con Mayor Productividad", heading_style))
                        img = RLImage(img_path, width=5.0*inch, height=3.5*inch)
                        elements.append(img)
                        elements.append(Spacer(1, 0.3*inch))

            # Descripciones generales
            if 'descripciones_general' in data and data['descripciones_general']:
                elements.append(PageBreak())
                elements.append(Paragraph("Análisis de Motivos por Estado", heading_style))
                
                for estado, df_desc in data['descripciones_general'].items():
                    if not df_desc.empty:
                        elements.append(Paragraph(f"Motivos - {estado.title()}", styles['Heading3']))
                        
                        df_desc_limited = df_desc.head(10).copy()
                        table_data = [[ReportService._format_header_text('Motivo'), ReportService._format_header_text('Total')]]
                        for _, row in df_desc_limited.iterrows():
                            motivo_text = Paragraph(str(row['motivo']), styles['Normal'])
                            table_data.append([motivo_text, row['total']])
                        
                        table = Table(table_data, colWidths=[4*inch, 1.5*inch])
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 9),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ]))
                        elements.append(table)
                        elements.append(Spacer(1, 0.2*inch))

            # Descripciones por trámite
            if 'descripciones_tramites' in data and data['descripciones_tramites']:
                elements.append(PageBreak())
                elements.append(Paragraph("Motivos por Trámite", heading_style))
                
                for estado, df_desc in data['descripciones_tramites'].items():
                    if not df_desc.empty:
                        elements.append(Paragraph(f"Estado: {estado.title()}", styles['Heading3']))
                        
                        df_desc_limited = df_desc.head(15).copy()
                        columnas_formateadas = [ReportService._format_header_text(col) for col in df_desc_limited.columns]
                        table_data = [columnas_formateadas]
                        
                        for _, row in df_desc_limited.iterrows():
                            tramite_text = Paragraph(str(row['tramite']), styles['Normal'])
                            motivo_text = Paragraph(str(row['motivo']), styles['Normal'])
                            table_data.append([tramite_text, motivo_text, row['total']])
                        
                        table = Table(table_data, colWidths=[2*inch, 2.5*inch, 1*inch])
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 9),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ]))
                        elements.append(table)
                        elements.append(Spacer(1, 0.2*inch))

            # Descripciones por usuario
            if 'descripciones_usuarios' in data and data['descripciones_usuarios']:
                elements.append(PageBreak())
                elements.append(Paragraph("Motivos por Usuario", heading_style))
                
                for estado, df_desc in data['descripciones_usuarios'].items():
                    if not df_desc.empty:
                        elements.append(Paragraph(f"Estado: {estado.title()}", styles['Heading3']))
                        
                        df_desc_limited = df_desc.head(15).copy()
                        columnas_formateadas = [ReportService._format_header_text(col) for col in df_desc_limited.columns]
                        table_data = [columnas_formateadas]
                        
                        for _, row in df_desc_limited.iterrows():
                            usuario_text = Paragraph(str(row['usuario']), styles['Normal'])
                            motivo_text = Paragraph(str(row['motivo']), styles['Normal'])
                            table_data.append([usuario_text, motivo_text, row['total']])
                        
                        table = Table(table_data, colWidths=[2*inch, 2.5*inch, 1*inch])
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 9),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ]))
                        elements.append(table)
                        elements.append(Spacer(1, 0.2*inch))

            # Desglose por Empleado
            if 'usuarios' in data and not data['usuarios'].empty:
                tramites_por_emp = ReportService._tramites_por_empleado(filtros)
                if tramites_por_emp:
                    elements.append(PageBreak())
                    elements.append(Paragraph("Desglose de Trámites por Empleado", heading_style))
                    
                    for empleado, df_emp in tramites_por_emp.items():
                        emp_display = (empleado[:40] + '...') if len(empleado) > 40 else empleado
                        elements.append(Paragraph(f"Empleado: {emp_display}", styles['Heading3']))
                        
                        df_emp_limited = df_emp.head(10).copy()
                        table_data = [[ReportService._format_header_text('Trámite'), ReportService._format_header_text('Atenciones')]]
                        for _, row in df_emp_limited.iterrows():
                            tramite_text = Paragraph(row['tramite'], styles['Normal'])
                            table_data.append([tramite_text, row['total']])
                        
                        table = Table(table_data, colWidths=[4*inch, 1.5*inch])
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 9),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ]))
                        elements.append(table)
                        elements.append(Spacer(1, 0.2*inch))

            # Horas Pico
            if 'horas_pico' in data and not data['horas_pico'].empty and data['horas_pico']['total'].sum() > 0:
                elements.append(PageBreak())
                elements.append(Paragraph("Distribución de Atenciones por Hora", heading_style))
                img_path = ReportService._crear_grafica_horas_pico(data['horas_pico'])
                if img_path:
                    temp_image_paths.append(img_path)
                    img = RLImage(img_path, width=5.5*inch, height=3.0*inch)
                    elements.append(img)
                    elements.append(Spacer(1, 0.3*inch))

            # Patrón Semanal
            if 'horas_pico_dia' in data and not data['horas_pico_dia'].empty and data['horas_pico_dia']['total'].sum() > 0:
                elements.append(PageBreak())
                elements.append(Paragraph("Patrón Semanal de Atenciones", heading_style))
                img_path = ReportService._crear_heatmap_horas_dia(data['horas_pico_dia'])
                if img_path:
                    temp_image_paths.append(img_path)
                    img = RLImage(img_path, width=5.5*inch, height=3.5*inch)
                    elements.append(img)

            doc.build(elements)

        finally:
            for path in temp_image_paths:
                try:
                    os.unlink(path)
                except Exception as e:
                    print(f"Advertencia: No se pudo eliminar archivo temporal {path}: {e}")

        buffer.seek(0)
        return buffer

    @staticmethod
    def _crear_grafica_estados(resumen):
        try:
            estados = ['Atendidas', 'Reasignadas', 'Canceladas']
            valores = [
                resumen.get('finalizadas', 0),
                resumen.get('reasignadas', 0),
                resumen.get('canceladas', 0)
            ]
            if sum(valores) == 0:
                return None
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                fig, ax = plt.subplots(figsize=(6, 6))
                colors_pie = ['#2ecc71', '#f39c12', '#e74c3c']
                ax.pie(valores, labels=estados, autopct='%1.1f%%', startangle=90, colors=colors_pie)
                ax.set_title('Distribución de Estados de Atenciones', fontsize=12, fontweight='bold', pad=20)
                plt.tight_layout(pad=2.0)
                plt.savefig(tmp.name, dpi=150, bbox_inches='tight')
                plt.close()
                return tmp.name
        except Exception as e:
            print(f"Error creando gráfica de estados: {e}")
            return None

    @staticmethod
    def _crear_grafica_top_tramites(df):
        try:
            if df.empty or df['total_atenciones'].sum() == 0:
                return None
            df_sorted = df.nlargest(10, 'total_atenciones').copy()
            df_sorted['nombre_tramite'] = df_sorted['nombre_tramite'].astype(str).apply(
                lambda x: (x[:55] + '...') if len(x) > 55 else x
            )
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                fig, ax = plt.subplots(figsize=(7, 5))
                bars = ax.barh(df_sorted['nombre_tramite'], df_sorted['total_atenciones'], color='#3498db')
                ax.set_xlabel('Total de Atenciones', fontsize=10)
                ax.set_title('Trámites con Mayor Demanda', fontsize=12, fontweight='bold', pad=15)
                ax.invert_yaxis()
                ax.tick_params(axis='y', labelsize=8)
                ax.set_xlim(0, df_sorted['total_atenciones'].max() * 1.1)
                for bar in bars:
                    width = bar.get_width()
                    ax.text(width + 0.5, bar.get_y() + bar.get_height()/2, str(int(width)),
                            va='center', ha='left', fontsize=8)
                plt.tight_layout(pad=2.0)
                plt.savefig(tmp.name, dpi=150, bbox_inches='tight')
                plt.close()
                return tmp.name
        except Exception as e:
            print(f"Error creando gráfica de trámites: {e}")
            return None

    @staticmethod
    def _crear_grafica_top_empleados(df):
        try:
            if df.empty or df['total_atenciones'].sum() == 0:
                return None
            df_sorted = df.nlargest(10, 'total_atenciones').copy()
            df_sorted['nombre_completo'] = df_sorted['nombre_completo'].astype(str).apply(
                lambda x: (x[:35] + '...') if len(x) > 35 else x
            )
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                fig, ax = plt.subplots(figsize=(7, 5))
                bars = ax.barh(df_sorted['nombre_completo'], df_sorted['total_atenciones'], color='#9b59b6')
                ax.set_xlabel('Total de Atenciones', fontsize=10)
                ax.set_title('Empleados con Mayor Productividad', fontsize=12, fontweight='bold', pad=15)
                ax.invert_yaxis()
                ax.tick_params(axis='y', labelsize=8)
                ax.set_xlim(0, df_sorted['total_atenciones'].max() * 1.1)
                for bar in bars:
                    width = bar.get_width()
                    ax.text(width + 0.5, bar.get_y() + bar.get_height()/2, str(int(width)),
                            va='center', ha='left', fontsize=8)
                plt.tight_layout(pad=2.0)
                plt.savefig(tmp.name, dpi=150, bbox_inches='tight')
                plt.close()
                return tmp.name
        except Exception as e:
            print(f"Error creando gráfica de empleados: {e}")
            return None

    @staticmethod
    def _crear_grafica_horas_pico(df):
        try:
            if df.empty or df['total'].sum() == 0:
                return None
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                fig, ax = plt.subplots(figsize=(8, 4))
                ax.plot(df['hora'], df['total'], marker='o', linewidth=2, markersize=5, color='#e74c3c')
                ax.fill_between(df['hora'], df['total'], alpha=0.2, color='#e74c3c')
                ax.set_xlabel('Hora del Día', fontsize=10)
                ax.set_ylabel('Número de Atenciones', fontsize=10)
                ax.set_title('Distribución de Atenciones por Hora', fontsize=12, fontweight='bold', pad=15)
                ax.grid(True, alpha=0.3)
                min_hora = df['hora'].min()
                max_hora = df['hora'].max()
                ax.set_xlim(max(0, min_hora - 1), min(23, max_hora + 1))
                ax.set_xticks(range(int(min_hora), int(max_hora) + 1))
                ax.tick_params(axis='x', labelsize=8)
                ax.tick_params(axis='y', labelsize=8)
                plt.tight_layout(pad=2.0)
                plt.savefig(tmp.name, dpi=150, bbox_inches='tight')
                plt.close()
                return tmp.name
        except Exception as e:
            print(f"Error creando gráfica de horas pico: {e}")
            return None

    @staticmethod
    def _crear_heatmap_horas_dia(df):
        try:
            if df.empty or df['total'].sum() == 0:
                return None
            pivot_data = df.pivot_table(values='total', index='hora', columns='dia_nombre', fill_value=0)
            dias_orden = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
            pivot_data = pivot_data.reindex(columns=[d for d in dias_orden if d in pivot_data.columns], fill_value=0)
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                fig, ax = plt.subplots(figsize=(8, 5))
                sns.heatmap(pivot_data, annot=True, fmt='g', cmap='YlOrRd', ax=ax,
                            cbar_kws={'label': 'Atenciones'}, annot_kws={"size": 8})
                ax.set_xlabel('Día de la Semana', fontsize=10)
                ax.set_ylabel('Hora del Día', fontsize=10)
                ax.set_title('Patrón Semanal de Atenciones', fontsize=12, fontweight='bold', pad=15)
                ax.tick_params(axis='x', labelsize=8, rotation=30)
                ax.tick_params(axis='y', labelsize=8)
                plt.tight_layout(pad=2.0)
                plt.savefig(tmp.name, dpi=150, bbox_inches='tight')
                plt.close()
                return tmp.name
        except Exception as e:
            print(f"Error creando heatmap: {e}")
            return None